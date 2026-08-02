import logging
from io import BytesIO
from typing import List, Dict, Any
import openpyxl

logger = logging.getLogger(__name__)


class ExcelHeaderService:
    def extract_headers(self, file_content: bytes, file_name: str) -> List[Dict[str, Any]]:
        """Parses an Excel spreadsheet in bytes format and extracts headers from all sheets.
        
        Returns a list of dictionaries, one per sheet:
        [
            {
                "file_name": "PSUR Current.xlsx",
                "sheet_name": "Cases",
                "headers": [
                    {"column_index": 1, "original_header": "Subject"},
                    {"column_index": 2, "original_header": "Years"},
                    ...
                ]
            }
        ]
        """
        logger.info(f"Extracting headers from Excel file: {file_name}")
        try:
            # Load workbook from bytes in read-only mode for performance
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            results = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = []
                
                # Retrieve the first row
                first_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
                if first_row:
                    for idx, cell_value in enumerate(first_row):
                        if cell_value is not None:
                            header_str = str(cell_value).strip()
                            if header_str:
                                headers.append({
                                    "column_index": idx + 1,  # 1-indexed column position
                                    "original_header": header_str
                                })
                
                results.append({
                    "file_name": file_name,
                    "sheet_name": sheet_name,
                    "headers": headers
                })
                
            return results
        except Exception as e:
            logger.error(f"Failed to read headers from Excel file '{file_name}': {e}")
            raise ValueError(f"Failed to process Excel file structure: {str(e)}")


excel_header_service = ExcelHeaderService()
