import logging
from io import BytesIO
from typing import List, Dict, Any
import openpyxl

logger = logging.getLogger(__name__)


class ExcelDataService:
    def extract_rows(
        self, 
        file_content: bytes, 
        file_name: str, 
        expected_file_type: str,
        mappings: Dict[str, str]  # {uploaded_header: mapped_logical_field}
    ) -> List[Dict[str, Any]]:
        """Parses cell data from an Excel spreadsheet and converts row headers to logical fields.
        
        Appends source provenance details ('_source') to each row record.
        """
        logger.info(f"Extracting rows from file '{file_name}' using mappings: {mappings}")
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=True)
            extracted_rows = []
            
            # Convert mappings to lowercase keys for case-insensitive lookup
            clean_mappings = {k.strip().lower(): v for k, v in mappings.items()}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get headers row
                rows_iter = sheet.iter_rows(values_only=True)
                first_row = next(rows_iter, None)
                if not first_row:
                    continue
                
                # Map column index (0-based) to logical field name
                col_to_field = {}
                for idx, cell_val in enumerate(first_row):
                    if cell_val is not None:
                        clean_hdr = str(cell_val).strip().lower()
                        if clean_hdr in clean_mappings:
                            col_to_field[idx] = clean_mappings[clean_hdr]

                if not col_to_field:
                    logger.warning(f"No mapped headers found in sheet '{sheet_name}' of file '{file_name}'.")
                    continue

                # Read subsequent rows
                for r_idx, row_values in enumerate(rows_iter, start=2):
                    # Check if row is completely empty
                    if all(val is None or str(val).strip() == "" for val in row_values):
                        continue
                        
                    row_data = {}
                    for c_idx, val in enumerate(row_values):
                        if c_idx in col_to_field:
                            field_name = col_to_field[c_idx]
                            row_data[field_name] = val
                    
                    # Attach source provenance metadata
                    row_data["_source"] = {
                        "file": file_name,
                        "sheet": sheet_name,
                        "row": r_idx
                    }
                    
                    extracted_rows.append(row_data)
                    
            logger.info(f"Extracted {len(extracted_rows)} raw rows from '{file_name}'")
            return extracted_rows
        except Exception as e:
            logger.error(f"Failed to extract rows from Excel file '{file_name}': {e}")
            raise ValueError(f"Failed to process data rows in Excel: {str(e)}")


excel_data_service = ExcelDataService()
